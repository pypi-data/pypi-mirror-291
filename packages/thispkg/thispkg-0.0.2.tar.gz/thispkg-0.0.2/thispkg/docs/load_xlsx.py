import json
from datetime import date, datetime, time, timedelta
from decimal import Decimal

from openpyxl import load_workbook

from ..base64 import base64
from ._base import RawDoc


class JsonEncoder(json.JSONEncoder):
    def default(self, o: object) -> object:
        if isinstance(o, datetime):
            return o.isoformat()
        if isinstance(o, date):
            return o.isoformat()
        if isinstance(o, time):
            return o.isoformat()
        if isinstance(o, timedelta):
            return o.total_seconds()
        if isinstance(o, Decimal):
            return str(o)
        return super().default(o)


class ExcelLoader(RawDoc):
    def extract_text(self):
        with self.get_file() as file:
            wb = load_workbook(filename=file, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            data_dict = {
                                "row": cell.row,
                                "column": cell.column,
                                "value": cell.value,
                                "sheet": sheet_name,
                            }
                            yield json.dumps(data_dict, cls=JsonEncoder)

    def extract_image(self):
        with self.get_file() as file:
            wb = load_workbook(filename=file)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for image in sheet._images:  # type: ignore
                    yield base64.b64encode(image._data()).decode()  # type: ignore
