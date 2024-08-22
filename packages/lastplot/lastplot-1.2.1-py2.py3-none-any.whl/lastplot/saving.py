import contextlib
import itertools

import pandas as pd
from blessed import Terminal
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def color_df(df_final, sheet, output_path, output_file):

    # Generate unique region names
    unique_regions = df_final["Regions"].unique()

    # Generate a list of colors (repeat colors if there are more regions than colors)
    colors = ["EA9EBF", "D0A0EC", "B4DFFF", "FFEEBC", "DCEDC1"]
    color_cycle = itertools.cycle(colors)

    # Create a dictionary to map regions to colors
    region_fill_colors = {
        region: PatternFill(start_color=color, end_color=color, fill_type="solid")
        for region, color in zip(unique_regions, color_cycle)
    }

    with open_excel(output_path + "/output/" + output_file + ".xlsx") as wb:
        ws = wb[sheet]

        # Iterate through rows and apply fill colors based on 'Regions' column
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column
        ):
            region_value = row[0].value
            region_fill = region_fill_colors.get(region_value, None)

            if region_fill:
                for cell in row:
                    cell.fill = region_fill


def save_values(df_final, output_path, output_file):
    df_save = df_final.pivot_table(
        index=["Regions", "Mouse ID", "Genotype"],
        columns=["Lipids", "Lipid Class"],
        values=["Values", "Log10 Values"],
    )
    df_save.reset_index(inplace=True)

    # Save the eliminated lipids and the normalized data with the Z Scores
    with write_excel(
        output_path + "/output/" + output_file + ".xlsx", engine="openpyxl", mode="a"
    ) as writer:
        df_save.to_excel(writer, sheet_name="Values and Transformed Values")

    color_df(
        df_final,
        sheet="Values and Transformed Values",
        output_path=output_path,
        output_file=output_file,
    )


def save_zscores(df_final, output_path, output_file):
    df_save = df_final.pivot_table(
        index=["Regions", "Mouse ID", "Genotype"],
        columns=["Lipids", "Lipid Class"],
        values=["Z_Scores"],
    )
    df_save.reset_index(inplace=True)

    df_save2 = df_final.pivot_table(
        index=["Regions", "Mouse ID", "Genotype"],
        columns=["Lipid Class"],
        values=["Average Z_Scores"],
    )
    df_save2.reset_index(inplace=True)

    # Save the eliminated lipids and the normalized data with the Z Scores
    with write_excel(
        output_path + "/output/" + output_file + ".xlsx",
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay",
    ) as writer:
        df_save.to_excel(writer, sheet_name="Z_Scores", startrow=0, startcol=0)
        df_save2.to_excel(
            writer,
            sheet_name="Z_Scores",
            startrow=1,
            startcol=len(df_save.columns) + 1,
        )
        print("Saving to output file")

    color_df(
        df_final, sheet="Z_Scores", output_path=output_path, output_file=output_file
    )


def save_sheet(comment, sheet_name, output_path, output_file):
    path = output_path + "/output/" + output_file + ".xlsx"
    with open_excel(path) as wb:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)

        ws = wb[sheet_name]
        ws.append(comment)
        wb.save(path)


@contextlib.contextmanager
def open_excel(path: str):
    wb = load_workbook(path)
    try:
        yield wb
    except PermissionError:
        with Terminal() as term:
            yell_at_terminal(
                f"COULD NOT OPEN '{path}'!\n    CLOSE EXCEL FILE AND TRY AGAIN."
            )
    finally:
        wb.close()


@contextlib.contextmanager
def write_excel(path: str, **kwargs):
    try:
        with pd.ExcelWriter(path, **kwargs) as writer:
            try:
                yield writer
            except PermissionError:
                yell_at_terminal(
                    f"COULD NOT OPEN '{path}'!\n    CLOSE EXCEL FILE AND TRY AGAIN."
                )
    except PermissionError:
        yell_at_terminal(
            f"COULD NOT OPEN '{path}'!\n    CLOSE EXCEL FILE AND TRY AGAIN."
        )


def yell_at_terminal(message: str):
    from rich import print, box
    from rich.panel import Panel

    print("\n\n\n")
    print(
        Panel(
            f"Your operation FAILED! \n\n - {message}",
            title="ERROR - IMPORTANT - READ ME",
            box=box.DOUBLE_EDGE,
        )
    )

    exit()
